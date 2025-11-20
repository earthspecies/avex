#!/usr/bin/env python3
"""
create_posttrained_winrate_standalone.py
=======================================

Creates a standalone bar chart showing the benefit of post-training SSL backbones.
This is panel (a) from the post-trained win-rate analysis as a separate figure.

Usage:
    uv run python scripts/analysis/create_posttrained_winrate_standalone.py
"""

from __future__ import annotations

import datetime
from pathlib import Path
from typing import Dict, Tuple

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
from scipy.stats import pearsonr

# Configuration
EXCEL = Path("~/Downloads/results_with_birdmae.xlsx").expanduser()
OUT_BASE = Path("analysis")

# Post-training pairs: (post-trained model, base model)
POST_TRAINING_PAIRS = {
    "EAT-AS": ("sl-EAT-AS", "EAT-all"),  # Use EAT-all as base per user specification
    "EAT-bio": ("sl-EAT-bio", "EAT-all"),
    "EAT-all": ("sl-EAT-all", "EAT-all"),
    "BEATS-bio": ("sl-BEATS-bio", "BEATS (pretrained)"),
    "BEATS-all": ("sl-BEATS-all", "BEATS (pretrained)"),
}

# Benchmark-specific metrics (matching main analysis pipeline)
BENCHMARK_METRICS = {
    "BEANS Classification": ["Probe", "R-auc", "C-nmi"],
    "BEANS Detection": ["Probe", "R-auc"],  # No C-nmi for detection
    "BirdSet": ["Probe", "R-auc"],  # No C-nmi for detection
    "Individual ID": ["Probe", "R-auc"],
    "Vocal Repertoire": ["R-auc", "C-nmi"],  # No Probe for vocal repertoire
}

AGGREGATE_PREFIXES = {
    "BEANS Classification",
    "BEANS Detection",
    "BirdSet",
    "Individual ID",
    "Repertoire",
    "Vocal Repertoire",
}

# Benchmark groups for win-rate analysis
BENCHMARK_GROUPS = {
    "BEANS Classification": ["Watkins", "CBI", "HBDB", "BATS", "Dogs", "ESC-50"],
    "BEANS Detection": ["enabirds", "rfcx", "hiceas", "gibbons", "dcase"],
    "BirdSet": ["POW", "PER", "NES", "NBP", "HSN", "SNE", "UHH"],
    "Individual ID": [
        "chiffchaff-cross",
        "littleowls-cross",
        "pipit-cross",
        "macaques",
    ],
    "Vocal Repertoire": [
        "zebrafinch-je-call",
        "Giant_Otters",
        "Bengalese_Finch",
        "SRKW_Orca",
    ],
}

# Model configuration - including both new and existing models + post-trained
MODEL_GROUPS = {
    "New Supervised Models": ["EffNetB0-all", "EffNetB0-bio", "EffNetB0-AudioSet"],
    "New SSL Models": ["EAT-AS", "EAT-bio", "EAT-all"],
    "Post-trained SSL Models": [
        "sl-EAT-bio",
        "sl-EAT-all",
        "sl-BEATS-bio",
        "sl-BEATS-all",
        "BEATS-NatureLM-audio",
    ],
    "Existing Supervised Models": ["Perch", "BirdNet", "SurfPerch"],
    "Existing SSL Models": [
        "EAT-base (pretrained)",
        "EAT-base (SFT)",
        "BEATS (SFT)",
        "BEATS (pretrained)",
        "Bird-AVES-biox-base",
        "BirdMAE (pretrained)",
    ],
}

# Get all models as flat list
ALL_MODELS = []
for group_models in MODEL_GROUPS.values():
    ALL_MODELS.extend(group_models)


def load_df(path: Path) -> pd.DataFrame:
    """Flatten Excel into model × metric DataFrame.

    Returns
    -------
    pd.DataFrame
        Flattened dataframe with models as rows and metrics as columns.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers: list[str | None] = []
    for col in range(3, ws.max_column + 1):
        ds = ws.cell(row=3, column=col).value
        met = ws.cell(row=4, column=col).value
        headers.append(
            f"{str(ds).replace(' ', '_')}_{str(met).replace(' ', '_')}" if ds and met else None
        )
    rows = []
    for row in range(5, ws.max_row + 1):
        model = ws.cell(row=row, column=2).value
        if not model:
            continue
        rec: dict[str, float] = {"model": model}
        for idx, hdr in enumerate(headers, start=3):
            if hdr is None:
                continue
            val = ws.cell(row=row, column=idx).value
            try:
                rec[hdr] = float(val) if val not in (None, "", "N/A") else np.nan
            except Exception:
                rec[hdr] = np.nan
        rows.append(rec)
    return pd.DataFrame(rows).set_index("model")


def compute_benchmark_win_rates(
    df: pd.DataFrame, post_trained: str, base: str
) -> Dict[str, Tuple[int, int, float]]:
    """Compute win-rates by benchmark group.

    Uses appropriate metrics for each benchmark.

    Returns
    -------
    Dict[str, Tuple[int, int, float]]
        Dictionary mapping benchmark names to (wins, total, win_rate) tuples.
    """
    if post_trained not in df.index or base not in df.index:
        return {}

    g = pd.to_numeric(df.loc[post_trained], errors="coerce")
    b = pd.to_numeric(df.loc[base], errors="coerce")
    imp = (g - b) / b.replace(0, np.nan) * 100

    benchmark_wins = {}

    for benchmark_name, datasets in BENCHMARK_GROUPS.items():
        # Get metrics for this specific benchmark
        valid_metrics = BENCHMARK_METRICS.get(benchmark_name, [])

        # Collect improvement values for this benchmark
        benchmark_improvements = []

        for metric in valid_metrics:
            if benchmark_name == "Individual ID" and metric == "R-auc":
                # Special handling: prefer R-cross-auc over R-auc when available
                # (matching main analysis)
                for dataset in datasets:
                    cross_key = f"{dataset}_R-cross-auc"
                    regular_key = f"{dataset}_R-auc"

                    if cross_key in imp.index and pd.notna(imp[cross_key]):
                        # Use cross-auc when available (more robust)
                        benchmark_improvements.append(imp[cross_key])
                    elif regular_key in imp.index and pd.notna(imp[regular_key]):
                        # Fall back to regular R-auc
                        benchmark_improvements.append(imp[regular_key])
            else:
                # Standard handling for all other benchmark/metric combinations
                for dataset in datasets:
                    col_name = f"{dataset}_{metric}"
                    if col_name in imp.index and pd.notna(imp[col_name]):
                        benchmark_improvements.append(imp[col_name])

        if benchmark_improvements:
            # Convert to series for win_rate calculation
            benchmark_series = pd.Series(benchmark_improvements)
            wins = int((benchmark_series > 0).sum())
            total = int(benchmark_series.size)
            avg_imp = benchmark_series.mean()
            benchmark_wins[benchmark_name] = (wins, total, avg_imp)

    return benchmark_wins


def aggregate_win_rates_across_models(
    df: pd.DataFrame,
) -> Dict[str, Tuple[int, int, float]]:
    """Compute aggregated win-rates across all post-training pairs.

    Computes win-rates for each benchmark.

    Returns
    -------
    Dict[str, Tuple[int, int, float]]
        Dictionary mapping benchmark names to (wins, total, win_rate) tuples.
    """

    aggregated_wins = {}

    # Initialize counters for each benchmark
    for benchmark_name in BENCHMARK_GROUPS.keys():
        total_wins = 0
        total_comparisons = 0
        all_improvements = []

        # Compute win-rates for each post-training pair
        for _model_name, (post_trained, base) in POST_TRAINING_PAIRS.items():
            benchmark_wins = compute_benchmark_win_rates(df, post_trained, base)

            if benchmark_name in benchmark_wins:
                wins, total, avg_imp = benchmark_wins[benchmark_name]
                total_wins += wins
                total_comparisons += total

                # Collect individual improvements for overall average
                g = pd.to_numeric(df.loc[post_trained], errors="coerce")
                b = pd.to_numeric(df.loc[base], errors="coerce")
                imp = (g - b) / b.replace(0, np.nan) * 100

                # Get improvements for this benchmark
                # (same logic as compute_benchmark_win_rates)
                valid_metrics = BENCHMARK_METRICS.get(benchmark_name, [])
                for metric in valid_metrics:
                    if benchmark_name == "Individual ID" and metric == "R-auc":
                        for dataset in BENCHMARK_GROUPS[benchmark_name]:
                            cross_key = f"{dataset}_R-cross-auc"
                            regular_key = f"{dataset}_R-auc"

                            if cross_key in imp.index and pd.notna(imp[cross_key]):
                                all_improvements.append(imp[cross_key])
                            elif regular_key in imp.index and pd.notna(imp[regular_key]):
                                all_improvements.append(imp[regular_key])
                    else:
                        for dataset in BENCHMARK_GROUPS[benchmark_name]:
                            col_name = f"{dataset}_{metric}"
                            if col_name in imp.index and pd.notna(imp[col_name]):
                                all_improvements.append(imp[col_name])

        if total_comparisons > 0:
            avg_improvement = np.mean(all_improvements) if all_improvements else 0.0
            aggregated_wins[benchmark_name] = (
                total_wins,
                total_comparisons,
                avg_improvement,
            )

    return aggregated_wins


def compute_beans_averages(df: pd.DataFrame, exclude_esc50: bool = False) -> pd.DataFrame:
    """Compute BEANS Classification and Detection aggregated averages.

    Parameters
    ----------
    df : pd.DataFrame
        Input dataframe with model metrics.
    exclude_esc50 : bool, default False
        If True, exclude ESC-50 from BEANS Classification aggregation.

    Returns
    -------
    pd.DataFrame
        Dataframe with computed BEANS averages.
    """
    result = pd.DataFrame(index=df.index)

    # BEANS Classification datasets
    if exclude_esc50:
        beans_classification = ["Watkins", "CBI", "HBDB", "BATS", "Dogs"]
    else:
        beans_classification = ["Watkins", "CBI", "HBDB", "BATS", "Dogs", "ESC-50"]

    # BEANS Detection datasets
    beans_detection = ["enabirds", "rfcx", "hiceas", "gibbons", "dcase"]

    # Compute R-auc average for BEANS Classification
    metric = "R-auc"
    keys = [f"{ds}_{metric}" for ds in beans_classification if f"{ds}_{metric}" in df.columns]
    if keys:
        values = df[keys].apply(pd.to_numeric, errors="coerce")
        avg = values.mean(axis=1, skipna=True)
        result["BEANS Classification_R-auc"] = avg

    # Compute R-auc average for BEANS Detection
    keys = [f"{ds}_{metric}" for ds in beans_detection if f"{ds}_{metric}" in df.columns]
    if keys:
        values = df[keys].apply(pd.to_numeric, errors="coerce")
        avg = values.mean(axis=1, skipna=True)
        result["BEANS Detection_R-auc"] = avg

    return result


def compute_birdset_averages(df: pd.DataFrame) -> pd.DataFrame:
    """Compute BirdSet aggregated averages.

    Returns
    -------
    pd.DataFrame
        Dataframe with computed BirdSet averages.
    """
    result = pd.DataFrame(index=df.index)

    # BirdSet datasets
    birdset_datasets = ["POW", "PER", "NES", "NBP", "HSN", "SNE", "UHH"]

    # Compute R-auc average for BirdSet
    metric = "R-auc"
    keys = [f"{ds}_{metric}" for ds in birdset_datasets if f"{ds}_{metric}" in df.columns]
    if keys:
        values = df[keys].apply(pd.to_numeric, errors="coerce")
        # Exclude BirdNet from BirdSet averages (following main analysis)
        if "BirdNet" in result.index:
            # BirdNet should be excluded from BirdSet, but we'll compute it anyway
            # and let the user decide if they want to filter it
            pass
        avg = values.mean(axis=1, skipna=True)
        result["BirdSet_R-auc"] = avg

    return result


def create_cbi_birdset_scatter_plot(df: pd.DataFrame, output_dir: Path) -> None:
    """Create CBI vs BirdSet aggregated results scatter plot.

    Similar to the CBI vs BEANS Detection plot but with BirdSet on Y-axis.
    """
    # Compute BirdSet aggregated results
    birdset_averages = compute_birdset_averages(df)

    # Filter to all models that exist in our data
    available_models = [model for model in ALL_MODELS if model in birdset_averages.index]
    if not available_models:
        print("No models found in data!")
        return

    birdset_data = birdset_averages.loc[available_models]

    # Add model group information for coloring
    model_groups = {}
    for group_name, models in MODEL_GROUPS.items():
        for model in models:
            if model in available_models:
                model_groups[model] = group_name

    birdset_data["model_group"] = [model_groups.get(model, "Other") for model in birdset_data.index]

    # Metrics to plot - CBI dataset specifically vs BirdSet aggregated
    x_metric = "CBI_R-auc"  # CBI dataset ROC AUC specifically
    y_metric = "BirdSet_R-auc"  # Aggregated BirdSet ROC AUC

    # Check if metrics exist - if CBI_R-auc doesn't exist, try to get it from raw data
    if x_metric not in birdset_data.columns:
        # Try to get CBI directly from the original dataframe
        if "CBI_R-auc" in df.columns:
            # Add CBI data to birdset_data
            for model in available_models:
                if model in df.index:
                    birdset_data.loc[model, x_metric] = df.loc[model, "CBI_R-auc"]
        else:
            print(f"CBI dataset not found. Available columns: {list(df.columns)}")
            return

    if y_metric not in birdset_data.columns:
        print(f"Missing BirdSet metric. Available: {list(birdset_data.columns)}")
        return

    # Remove rows with NaN values
    mask = ~(birdset_data[x_metric].isna() | birdset_data[y_metric].isna())
    plot_data = birdset_data[mask]

    if len(plot_data) == 0:
        print("No valid CBI vs BirdSet data for plotting")
        return

    # === CREATE PLOT ===
    fig, ax = plt.subplots(1, 1, figsize=(10, 8))

    # Colorblind-safe colors: Blue for SSL, Cyan for SL
    ssl_color = "#00738B"  # Blue 4 for SSL
    sl_color = "#1ADCCF"  # Cyan 1 for Supervised Learning

    # Markers: circle for existing, X for new
    existing_marker = "o"  # Circle
    new_marker = "X"  # Cross

    # Define groups and their properties
    group_properties = {
        "New Supervised Models": {
            "color": sl_color,
            "marker": new_marker,
            "label": "New SL",
        },
        "Existing Supervised Models": {
            "color": sl_color,
            "marker": existing_marker,
            "label": "Existing SL",
        },
        "New SSL Models": {
            "color": ssl_color,
            "marker": new_marker,
            "label": "New SSL",
        },
        "Existing SSL Models": {
            "color": ssl_color,
            "marker": existing_marker,
            "label": "Existing SSL",
        },
        "Post-trained SSL Models": {
            "color": "black",
            "marker": new_marker,
            "label": "Post-trained SSL",
        },
    }

    # Plot each group with its factorial properties
    plotted_groups = []
    label_to_group = {}  # Map legend labels to group names
    for group_name, props in group_properties.items():
        group_data = plot_data[plot_data["model_group"] == group_name]
        if not group_data.empty:
            ax.scatter(
                group_data[x_metric],
                group_data[y_metric],
                label=props["label"],
                alpha=0.8,
                s=100,
                edgecolors="k",
                linewidths=0.5,
                color=props["color"],
                marker=props["marker"],
            )
            plotted_groups.append(group_name)
            label_to_group[props["label"]] = group_name

    # Add model labels with better positioning
    # Color labels to match marker colors (blue for SSL, orange for SL)
    for idx, row in plot_data.iterrows():
        if not pd.isna(row[x_metric]) and not pd.isna(row[y_metric]):
            label = idx
            if "Bird-AVES-biox-base" in label:
                label = "Bird-AVES"

            # Determine label color based on model group
            model_group = row.get("model_group", "Other")
            if model_group in ["New SSL Models", "Existing SSL Models"]:
                label_color = ssl_color
            elif model_group in ["New Supervised Models", "Existing Supervised Models"]:
                label_color = sl_color
            else:
                label_color = "black"  # Post-trained SSL and other groups use black

            ax.annotate(
                label,
                (row[x_metric], row[y_metric]),
                fontsize=15,
                alpha=0.8,
                ha="center",
                xytext=(2, 2),
                textcoords="offset points",
                color=label_color,
            )

    # Set axis limits with margins
    x_vals = plot_data[x_metric].dropna().to_numpy(dtype=float)
    y_vals = plot_data[y_metric].dropna().to_numpy(dtype=float)
    if x_vals.size and y_vals.size:
        x_min, x_max = x_vals.min(), x_vals.max()
        y_min, y_max = y_vals.min(), y_vals.max()
        x_margin = (x_max - x_min) * 0.08 if x_max > x_min else 0.01
        y_margin = (y_max - y_min) * 0.08 if y_max > y_min else 0.01
        ax.set_xlim(x_min - x_margin, x_max + x_margin)
        ax.set_ylim(y_min - y_margin, y_max + y_margin)

    # Calculate correlation
    corr_coef, p_value = pearsonr(plot_data[x_metric], plot_data[y_metric])

    ax.set_xlabel("CBI Dataset ROC-AUC", fontweight="bold", fontsize=12)
    ax.set_ylabel("Retrieval BirdSet R-AUC", fontweight="bold", fontsize=12)
    ax.set_title(
        "Performance on CBI vs BirdSet\nseparates by training paradigm",
        pad=15,
        fontweight="bold",
        fontsize=13,
    )
    ax.grid(True, alpha=0.25)

    # Create factorial legend: Blue/Orange = SSL/SL, Circle/X = Existing/New
    # Color the legend text to match the marker colors
    legend = ax.legend(
        loc="upper left",
        fontsize=15,
        title="Training Approach",
        title_fontsize=17,
        frameon=True,
        fancybox=True,
        shadow=True,
    )
    legend.get_title().set_fontweight("bold")

    # Color legend text to match marker colors
    for text in legend.get_texts():
        label_text = text.get_text()
        if label_text in label_to_group:
            group_name = label_to_group[label_text]
            if group_name in ["New SSL Models", "Existing SSL Models"]:
                text.set_color(ssl_color)
            elif group_name in ["New Supervised Models", "Existing Supervised Models"]:
                text.set_color(sl_color)
            else:
                text.set_color("black")  # Post-trained SSL and other groups use black

    plt.tight_layout()

    # Save the plot
    filename = "cbi_birdset_scatter.png"
    filepath = output_dir / filename
    plt.savefig(filepath, dpi=300, bbox_inches="tight")
    plt.close()

    print(f"✅ Created CBI vs BirdSet scatter plot: {filename}")


def create_beans_detection_classification_scatter_plot(
    df: pd.DataFrame, output_dir: Path, exclude_esc50: bool = False
) -> None:
    """Create BEANS Detection vs BEANS Classification scatter plot.

    Parameters
    ----------
    df : pd.DataFrame
        Input dataframe with model metrics.
    output_dir : Path
        Directory to save the plot.
    exclude_esc50 : bool, default False
        If True, exclude ESC-50 from BEANS Classification aggregation.
    """
    # Compute BEANS averages
    beans_averages = compute_beans_averages(df, exclude_esc50=exclude_esc50)

    # Filter to all models that exist in our data
    available_models = [model for model in ALL_MODELS if model in beans_averages.index]
    if not available_models:
        print("No models found in data!")
        return

    beans_data = beans_averages.loc[available_models]

    # Add model group information for coloring
    model_groups = {}
    for group_name, models in MODEL_GROUPS.items():
        for model in models:
            if model in available_models:
                model_groups[model] = group_name

    beans_data["model_group"] = [model_groups.get(model, "Other") for model in beans_data.index]

    # Metrics to plot
    x_metric = "BEANS Classification_R-auc"  # BEANS Classification aggregated
    y_metric = "BEANS Detection_R-auc"  # BEANS Detection aggregated

    if x_metric not in beans_data.columns:
        print(f"Missing BEANS Classification metric. Available: {list(beans_data.columns)}")
        return

    if y_metric not in beans_data.columns:
        print(f"Missing BEANS Detection metric. Available: {list(beans_data.columns)}")
        return

    # Remove rows with NaN values
    mask = ~(beans_data[x_metric].isna() | beans_data[y_metric].isna())
    plot_data = beans_data[mask]

    if len(plot_data) == 0:
        print("No valid BEANS Detection vs BEANS Classification data for plotting")
        return

    # === CREATE PLOT ===
    fig, ax = plt.subplots(1, 1, figsize=(10, 8))

    # Colorblind-safe colors: Blue for SSL, Cyan for SL
    ssl_color = "#00738B"  # Blue 4 for SSL
    sl_color = "#1ADCCF"  # Cyan 1 for Supervised Learning

    # Markers: circle for existing, X for new
    existing_marker = "o"  # Circle
    new_marker = "X"  # Cross

    # Define groups and their properties
    group_properties = {
        "New Supervised Models": {
            "color": sl_color,
            "marker": new_marker,
            "label": "New SL",
        },
        "Existing Supervised Models": {
            "color": sl_color,
            "marker": existing_marker,
            "label": "Existing SL",
        },
        "New SSL Models": {
            "color": ssl_color,
            "marker": new_marker,
            "label": "New SSL",
        },
        "Existing SSL Models": {
            "color": ssl_color,
            "marker": existing_marker,
            "label": "Existing SSL",
        },
        "Post-trained SSL Models": {
            "color": "black",
            "marker": new_marker,
            "label": "Post-trained SSL",
        },
    }

    # Plot each group with its factorial properties
    plotted_groups = []
    label_to_group = {}  # Map legend labels to group names
    for group_name, props in group_properties.items():
        group_data = plot_data[plot_data["model_group"] == group_name]
        if not group_data.empty:
            ax.scatter(
                group_data[x_metric],
                group_data[y_metric],
                label=props["label"],
                alpha=0.8,
                s=100,
                edgecolors="k",
                linewidths=0.5,
                color=props["color"],
                marker=props["marker"],
            )
            plotted_groups.append(group_name)
            label_to_group[props["label"]] = group_name

    # Add model labels with better positioning
    # Color labels to match marker colors (blue for SSL, orange for SL, black for others)
    for idx, row in plot_data.iterrows():
        if not pd.isna(row[x_metric]) and not pd.isna(row[y_metric]):
            label = idx
            if "Bird-AVES-biox-base" in label:
                label = "Bird-AVES"

            # Determine label color based on model group
            model_group = row.get("model_group", "Other")
            if model_group in ["New SSL Models", "Existing SSL Models"]:
                label_color = ssl_color
            elif model_group in ["New Supervised Models", "Existing Supervised Models"]:
                label_color = sl_color
            else:
                label_color = "black"  # Post-trained SSL and other groups use black

            ax.annotate(
                label,
                (row[x_metric], row[y_metric]),
                fontsize=15,
                alpha=0.8,
                ha="center",
                xytext=(2, 2),
                textcoords="offset points",
                color=label_color,
            )

    # Set axis limits with margins
    x_vals = plot_data[x_metric].dropna().to_numpy(dtype=float)
    y_vals = plot_data[y_metric].dropna().to_numpy(dtype=float)
    if x_vals.size and y_vals.size:
        x_min, x_max = x_vals.min(), x_vals.max()
        y_min, y_max = y_vals.min(), y_vals.max()
        x_margin = (x_max - x_min) * 0.08 if x_max > x_min else 0.01
        y_margin = (y_max - y_min) * 0.08 if y_max > y_min else 0.01
        ax.set_xlim(x_min - x_margin, x_max + x_margin)
        ax.set_ylim(y_min - y_margin, y_max + y_margin)

    # Calculate correlation
    corr_coef, p_value = pearsonr(plot_data[x_metric], plot_data[y_metric])

    esc50_text = " (w/o ESC-50)" if exclude_esc50 else ""
    ax.set_xlabel(
        f"Retrieval BEANS Classification R-AUC{esc50_text}", fontweight="bold", fontsize=12
    )
    ax.set_ylabel("Retrieval BEANS Detection R-AUC", fontweight="bold", fontsize=12)
    ax.set_title(
        f"Performance on BEANS Classification{esc50_text}"
        + "vs BEANS Detection\nseparates by training paradigm",
        pad=15,
        fontweight="bold",
        fontsize=13,
    )
    ax.grid(True, alpha=0.25)

    # Create factorial legend: Blue/Orange = SSL/SL, Circle/X = Existing/New
    # Color the legend text to match the marker colors
    legend = ax.legend(
        loc="upper left",
        fontsize=15,
        title="Training Approach",
        title_fontsize=17,
        frameon=True,
        fancybox=True,
        shadow=True,
    )
    legend.get_title().set_fontweight("bold")

    # Color legend text to match marker colors
    for text in legend.get_texts():
        label_text = text.get_text()
        if label_text in label_to_group:
            group_name = label_to_group[label_text]
            if group_name in ["New SSL Models", "Existing SSL Models"]:
                text.set_color(ssl_color)
            elif group_name in ["New Supervised Models", "Existing Supervised Models"]:
                text.set_color(sl_color)
            else:
                text.set_color("black")  # Post-trained SSL and other groups use black

    plt.tight_layout()

    # Save the plot
    suffix = "_noesc50" if exclude_esc50 else ""
    filename = f"beans_detection_classification_scatter{suffix}.png"
    filepath = output_dir / filename
    plt.savefig(filepath, dpi=300, bbox_inches="tight")
    plt.close()

    print(f"✅ Created BEANS Detection vs BEANS Classification scatter plot: {filename}")


def create_standalone_winrate_plot(df: pd.DataFrame, output_dir: Path) -> None:
    """Create standalone bar chart showing benefit of post-training SSL backbones."""

    # Compute aggregated win-rates across all post-training pairs
    aggregated_wins = aggregate_win_rates_across_models(df)

    if not aggregated_wins:
        print("No benchmark win-rate data available!")
        return

    # === CREATE PLOT ===
    fig, ax = plt.subplots(1, 1, figsize=(10, 7))

    # Extract data for plotting
    groups = list(aggregated_wins.keys())
    win_rates = [wins / total * 100 for wins, total, _ in aggregated_wins.values()]
    avg_gains = [avg_gain for _, _, avg_gain in aggregated_wins.values()]
    counts = [f"{wins}/{total}" for wins, total, _ in aggregated_wins.values()]

    # Shorten group names for better display
    display_groups = []
    for group in groups:
        if group == "BEANS Classification":
            display_groups.append("BEANS\nClassification")
        elif group == "BEANS Detection":
            display_groups.append("BEANS\nDetection")
        elif group == "Individual ID":
            display_groups.append("Individual\nID")
        elif group == "Vocal Repertoire":
            display_groups.append("Vocal\nRepertoire")
        else:
            display_groups.append(group)

    # Use ESP 5-tone palette for bars: Blue 4, Cyan 2, Blue 1, Cyan 1, Blue 2
    bars = ax.bar(
        display_groups,
        win_rates,
        color=["#00738B", "#04CDA0", "#C6DEE7", "#1ADCCF", "#98C6D2"],
        alpha=1.0,
        edgecolor="black",
        linewidth=1,
    )

    # Add labels on bars
    for bar, count, avg_gain in zip(bars, counts, avg_gains, strict=False):
        height = bar.get_height()
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            height + 1,
            f"{height:.1f}%\n{count}\n{avg_gain:+.1f}%",
            ha="center",
            va="bottom",
            fontweight="bold",
            fontsize=10,
        )

    ax.set_ylabel("Win-Rate (%)", fontweight="bold", fontsize=14)
    ax.set_title("Benefit of Post-training SSL Backbones", fontweight="bold", fontsize=16, pad=20)
    ax.set_ylim(0, 105)
    ax.grid(True, alpha=0.3, axis="y")

    # Improve layout
    plt.tight_layout()

    # Save the plot
    filename = "posttrained_winrate_standalone.png"
    filepath = output_dir / filename
    plt.savefig(filepath, dpi=300, bbox_inches="tight")
    plt.close()

    print(f"✅ Created standalone post-training win-rate plot: {filename}")


def main() -> None:
    """Main execution function."""
    print("🔬 Creating Standalone Post-training Win-Rate Analysis...")

    # Load data
    df = load_df(EXCEL)

    # Create output directory
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = OUT_BASE / f"{ts}_posttrained_winrate_standalone"
    out_dir.mkdir(parents=True, exist_ok=True)

    # Create the standalone plot
    create_standalone_winrate_plot(df, out_dir)

    # Create the CBI vs BirdSet scatter plot
    create_cbi_birdset_scatter_plot(df, out_dir)

    # Create the BEANS Detection vs BEANS Classification scatter plot (with ESC-50)
    create_beans_detection_classification_scatter_plot(df, out_dir, exclude_esc50=False)

    # Create the BEANS Detection vs BEANS Classification scatter plot (without ESC-50)
    create_beans_detection_classification_scatter_plot(df, out_dir, exclude_esc50=True)

    print(f"✅ Analysis complete! Results saved to: {out_dir}")


if __name__ == "__main__":
    main()
