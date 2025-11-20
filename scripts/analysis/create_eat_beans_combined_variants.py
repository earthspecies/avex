#!/usr/bin/env python3
"""
create_eat_beans_combined_variants.py
====================================

Creates three variants of the EAT + BEANs combined visualization:
1. By-benchmark-group win-rates + BEANs scatter using retrieval ROC AUC only
2. Both SSL and Supervised win-rates + BEANs scatter using retrieval ROC AUC only
3. Both SSL and Supervised win-rates + CBI vs BEANs Detection scatter

Usage:
    uv run python scripts/analysis/create_eat_beans_combined_variants.py
"""

from __future__ import annotations

import datetime
from pathlib import Path
from typing import Dict, Tuple

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd

# Configuration
EXCEL = Path("~/Downloads/results_with_birdmae.xlsx").expanduser()
OUT_BASE = Path("analysis")

# Win-rate analysis configuration
EAT_PAIRS = {
    "EAT (SSL)": ("EAT-all", "EAT-bio"),
}

# Add supervised pair for variant 2
BOTH_PAIRS = {
    "EAT (SSL)": ("EAT-all", "EAT-bio"),
    "EffNet (Supervised)": ("EffNetB0-all", "EffNetB0-bio"),
}

# Benchmark-specific metrics (matching main analysis pipeline)
BENCHMARK_METRICS = {
    "BEANS Classification": ["Probe", "R-auc", "C-nmi"],
    "BEANS Detection": ["Probe", "R-auc"],  # No C-nmi for detection
    "BirdSet": ["Probe", "R-auc"],  # No C-nmi for detection
    "Individual ID": ["Probe", "R-auc"],
    "Vocal Repertoire": ["R-auc", "C-nmi"],  # No Probe for vocal repertoire
}

AGGREGATE_PREFIXES = {
    "BEANS Classification",
    "BEANS Detection",
    "BirdSet",
    "Individual ID",
    "Repertoire",
    "Vocal Repertoire",
}

# Benchmark groups for win-rate analysis
BENCHMARK_GROUPS = {
    "BEANS Classification": ["Watkins", "CBI", "HBDB", "BATS", "Dogs", "ESC-50"],
    "BEANS Detection": ["enabirds", "rfcx", "hiceas", "gibbons", "dcase"],
    "BirdSet": ["POW", "PER", "NES", "NBP", "HSN", "SNE", "UHH"],
    "Individual ID": [
        "chiffchaff-cross",
        "littleowls-cross",
        "pipit-cross",
        "macaques",
    ],
    "Vocal Repertoire": [
        "zebrafinch-je-call",
        "Giant_Otters",
        "Bengalese_Finch",
        "SRKW_Orca",
    ],
}

# Model configuration - including both new and existing models + post-trained
MODEL_GROUPS = {
    "New Supervised Models": ["EffNetB0-all", "EffNetB0-bio", "EffNetB0-AudioSet"],
    "New SSL Models": ["EAT-AS", "EAT-bio", "EAT-all"],
    "Post-trained SSL Models": [
        "sl-EAT-bio",
        "sl-EAT-all",
        "sl-BEATS-bio",
        "sl-BEATS-all",
        "BEATS-NatureLM-audio",
    ],
    "Existing Supervised Models": ["Perch", "BirdNet", "SurfPerch"],
    "Existing SSL Models": [
        "EAT-base (pretrained)",
        "EAT-base (SFT)",
        "BEATS (SFT)",
        "BEATS (pretrained)",
        "Bird-AVES-biox-base",
        "BirdMAE (pretrained)",
    ],
}

# Get all models as flat list
ALL_MODELS = []
for group_models in MODEL_GROUPS.values():
    ALL_MODELS.extend(group_models)


def load_df(path: Path) -> pd.DataFrame:
    """Flatten Excel into model × metric DataFrame.

    Returns
    -------
    pd.DataFrame
        Flattened dataframe with models as rows and metrics as columns.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers: list[str | None] = []
    for col in range(3, ws.max_column + 1):
        ds = ws.cell(row=3, column=col).value
        met = ws.cell(row=4, column=col).value
        headers.append(
            f"{str(ds).replace(' ', '_')}_{str(met).replace(' ', '_')}" if ds and met else None
        )
    rows = []
    for row in range(5, ws.max_row + 1):
        model = ws.cell(row=row, column=2).value
        if not model:
            continue
        rec: dict[str, float] = {"model": model}
        for idx, hdr in enumerate(headers, start=3):
            if hdr is None:
                continue
            val = ws.cell(row=row, column=idx).value
            try:
                rec[hdr] = float(val) if val not in (None, "", "N/A") else np.nan
            except Exception:
                rec[hdr] = np.nan
        rows.append(rec)
    return pd.DataFrame(rows).set_index("model")


def improvement_series(df: pd.DataFrame, general: str, bio: str) -> pd.Series:
    """% improvement from bio -> general across metrics.

    Returns
    -------
    pd.Series
        Series with improvement percentages for each metric.
    """
    if general not in df.index or bio not in df.index:
        return pd.Series(dtype=float)
    g = pd.to_numeric(df.loc[general], errors="coerce")
    b = pd.to_numeric(df.loc[bio], errors="coerce")
    imp = (g - b) / b.replace(0, np.nan) * 100

    # Get all valid metrics from benchmark-specific definitions
    all_valid_metrics = set()
    for metrics in BENCHMARK_METRICS.values():
        all_valid_metrics.update(metrics)
    # Add R-cross-auc for Individual ID
    all_valid_metrics.add("R-cross-auc")

    # Keep only valid metrics & drop aggregate datasets
    filtered_cols = []
    for c in imp.index:
        metric_ok = c.split("_")[-1] in all_valid_metrics
        ds = "_".join(c.split("_")[:-1])
        agg_ok = not any(ds.startswith(prefix) for prefix in AGGREGATE_PREFIXES)
        if metric_ok and agg_ok:
            filtered_cols.append(c)
    imp = imp[filtered_cols]
    return imp.dropna()


def win_rate(series: pd.Series) -> tuple[int, int]:
    wins = int((series > 0).sum())
    total = int(series.size)
    return wins, total


def compute_benchmark_win_rates(
    df: pd.DataFrame, general: str, bio: str
) -> Dict[str, Tuple[int, int, float]]:
    """Compute win-rates by benchmark group.

    Uses appropriate metrics for each benchmark.

    Returns
    -------
    Dict[str, Tuple[int, int, float]]
        Dictionary mapping benchmark names to (wins, total, win_rate) tuples.
    """
    if general not in df.index or bio not in df.index:
        return {}

    g = pd.to_numeric(df.loc[general], errors="coerce")
    b = pd.to_numeric(df.loc[bio], errors="coerce")
    imp = (g - b) / b.replace(0, np.nan) * 100

    benchmark_wins = {}

    for benchmark_name, datasets in BENCHMARK_GROUPS.items():
        # Get metrics for this specific benchmark
        valid_metrics = BENCHMARK_METRICS.get(benchmark_name, [])

        # Collect improvement values for this benchmark
        benchmark_improvements = []

        for metric in valid_metrics:
            if benchmark_name == "Individual ID" and metric == "R-auc":
                # Special handling: prefer R-cross-auc over R-auc when available
                # (matching main analysis)
                for dataset in datasets:
                    cross_key = f"{dataset}_R-cross-auc"
                    regular_key = f"{dataset}_R-auc"

                    if cross_key in imp.index and pd.notna(imp[cross_key]):
                        # Use cross-auc when available (more robust)
                        benchmark_improvements.append(imp[cross_key])
                    elif regular_key in imp.index and pd.notna(imp[regular_key]):
                        # Fall back to regular R-auc
                        benchmark_improvements.append(imp[regular_key])
            else:
                # Standard handling for all other benchmark/metric combinations
                for dataset in datasets:
                    col_name = f"{dataset}_{metric}"
                    if col_name in imp.index and pd.notna(imp[col_name]):
                        benchmark_improvements.append(imp[col_name])

        if benchmark_improvements:
            # Convert to series for win_rate calculation
            benchmark_series = pd.Series(benchmark_improvements)
            wins, total = win_rate(benchmark_series)
            avg_imp = benchmark_series.mean()
            benchmark_wins[benchmark_name] = (wins, total, avg_imp)

    return benchmark_wins


def compute_benchmark_averages(df: pd.DataFrame) -> pd.DataFrame:
    """Compute BEANS Classification and Detection averages.

    Returns
    -------
    pd.DataFrame
        Dataframe with computed averages.
    """
    result = pd.DataFrame(index=df.index)

    # BEANS Classification datasets
    beans_classification = ["Watkins", "CBI", "HBDB", "BATS", "Dogs", "ESC-50"]
    # BEANS Detection datasets
    beans_detection = ["enabirds", "rfcx", "hiceas", "gibbons", "dcase"]

    benchmark_groups = {
        "BEANS Classification": beans_classification,
        "BEANS Detection": beans_detection,
    }

    for group_name, datasets in benchmark_groups.items():
        for metric in ["R-auc", "Probe"]:  # Core metrics
            keys = [f"{ds}_{metric}" for ds in datasets if f"{ds}_{metric}" in df.columns]
            if not keys:
                continue

            values = df[keys].apply(pd.to_numeric, errors="coerce")
            avg = values.mean(axis=1, skipna=True)
            result[f"{group_name}_{metric}"] = avg

    return result


def create_variant1_plot(df: pd.DataFrame, output_dir: Path) -> None:
    """Create variant 1: By-benchmark-group win-rates + BEANs scatter.

    Uses retrieval ROC AUC only (including post-trained models).
    """

    # === LEFT SUBPLOT: Win-rates by benchmark group ===
    benchmark_wins = compute_benchmark_win_rates(df, "EAT-all", "EAT-bio")

    if not benchmark_wins:
        print("No benchmark win-rate data available!")
        return

    # === RIGHT SUBPLOT: BEANs Classification vs Detection ===
    benchmark_averages = compute_benchmark_averages(df)

    # Filter to all models that exist in our data
    available_models = [model for model in ALL_MODELS if model in benchmark_averages.index]
    if not available_models:
        print("No models found in data!")
        return

    beans_data = benchmark_averages.loc[available_models]

    # Add model group information for coloring
    model_groups = {}
    for group_name, models in MODEL_GROUPS.items():
        for model in models:
            if model in available_models:
                model_groups[model] = group_name

    beans_data["model_group"] = [model_groups.get(model, "Other") for model in beans_data.index]

    # Metrics to plot - using only retrieval ROC AUC
    x_metric = "BEANS Classification_R-auc"  # Retrieval ROC AUC for BEANS Classification
    y_metric = "BEANS Detection_R-auc"  # Retrieval ROC AUC for BEANS Detection

    # Check if metrics exist
    if x_metric not in beans_data.columns or y_metric not in beans_data.columns:
        print(f"Missing BEANS metrics. Available: {list(beans_data.columns)}")
        return

    # Remove rows with NaN values
    mask = ~(beans_data[x_metric].isna() | beans_data[y_metric].isna())
    plot_data = beans_data[mask]

    if len(plot_data) == 0:
        print("No valid BEANS data for plotting")
        return

    # === CREATE PLOT ===
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 7))

    # LEFT: Win-rates by benchmark group
    groups = list(benchmark_wins.keys())
    win_rates = [wins / total * 100 for wins, total, _ in benchmark_wins.values()]
    avg_gains = [avg_gain for _, _, avg_gain in benchmark_wins.values()]
    counts = [f"{wins}/{total}" for wins, total, _ in benchmark_wins.values()]

    # Shorten group names for better display
    display_groups = []
    for group in groups:
        if group == "BEANS Classification":
            display_groups.append("BEANS\nClassification")
        elif group == "BEANS Detection":
            display_groups.append("BEANS\nDetection")
        elif group == "Individual ID":
            display_groups.append("Individual\nID")
        elif group == "Vocal Repertoire":
            display_groups.append("Vocal\nRepertoire")
        else:
            display_groups.append(group)

    # Updated palette for (a): Blue 4, Cyan 2, Blue 1, Cyan 1, Blue 2
    bars = ax1.bar(
        display_groups,
        win_rates,
        color=["#00738B", "#04CDA0", "#C6DEE7", "#1ADCCF", "#98C6D2"],
        alpha=1.0,
        edgecolor="black",
        linewidth=1,
    )

    # Add labels on bars
    for bar, count, avg_gain in zip(bars, counts, avg_gains, strict=False):
        height = bar.get_height()
        ax1.text(
            bar.get_x() + bar.get_width() / 2,
            height + 1,
            f"{height:.1f}%\n{count}\n{avg_gain:+.1f}%",
            ha="center",
            va="bottom",
            fontweight="bold",
            fontsize=14,
        )

    ax1.set_ylabel("Win-Rate (%)", fontweight="bold", fontsize=18)
    ax1.set_title(
        "Benefit of mixing general audio\nin pretraining",
        fontweight="bold",
        pad=15,
        fontsize=20,
    )
    ax1.set_ylim(0, 105)
    ax1.grid(True, alpha=0.3, axis="y")

    # Add panel label
    ax1.text(
        -0.1,
        1.05,
        "(a)",
        transform=ax1.transAxes,
        fontsize=24,
        fontweight="bold",
        va="top",
    )

    # RIGHT: BEANs scatter plot with factorial legend
    # High-contrast palette for (b) to match reference:
    #   SSL -> Blue 4 (#00738B)
    #   SL  -> Cyan 1 (#1ADCCF)
    ssl_color = "#00738B"  # Blue 4 for SSL
    sl_color = "#1ADCCF"  # Cyan 1 for Supervised Learning
    existing_marker = "o"  # Circle
    new_marker = "X"  # Cross
    group_properties = {
        "New Supervised Models": {
            "color": sl_color,
            "marker": new_marker,
            "label": "New SL",
        },
        "Existing Supervised Models": {
            "color": sl_color,
            "marker": existing_marker,
            "label": "Existing SL",
        },
        "New SSL Models": {
            "color": ssl_color,
            "marker": new_marker,
            "label": "New SSL",
        },
        "Existing SSL Models": {
            "color": ssl_color,
            "marker": existing_marker,
            "label": "Existing SSL",
        },
        "Post-trained SSL Models": {
            "color": "black",
            "marker": new_marker,
            "label": "Post-trained SSL",
        },
    }

    # Plot each group with its factorial properties
    plotted_groups = []
    label_to_group = {}  # Map legend labels to group names
    for group_name, props in group_properties.items():
        group_data = plot_data[plot_data["model_group"] == group_name]
        if not group_data.empty:
            ax2.scatter(
                group_data[x_metric],
                group_data[y_metric],
                label=props["label"],
                alpha=1.0,
                s=100,
                edgecolors="k",
                linewidths=0.5,
                color=props["color"],
                marker=props["marker"],
            )
            plotted_groups.append(group_name)
            label_to_group[props["label"]] = group_name

    # Add model labels with better positioning to avoid overflow
    # Color labels to match marker colors (blue for SSL, orange for SL, black for others)
    for idx, row in plot_data.iterrows():
        if not pd.isna(row[x_metric]) and not pd.isna(row[y_metric]):
            # Shorten Bird-AVES name to avoid overflow
            label = idx
            if "Bird-AVES-biox-base" in label:
                label = "Bird-AVES"

            # Determine label color based on model group
            model_group = row.get("model_group", "Other")
            if model_group in ["New SSL Models", "Existing SSL Models"]:
                label_color = ssl_color
            elif model_group in ["New Supervised Models", "Existing Supervised Models"]:
                label_color = sl_color
            else:
                label_color = "black"  # Post-trained SSL and other groups use black

            ax2.annotate(
                label,
                (row[x_metric], row[y_metric]),
                fontsize=15,
                alpha=0.8,
                ha="center",
                xytext=(2, 2),
                textcoords="offset points",
                color=label_color,
            )

    # Set axis limits with margins
    x_vals = plot_data[x_metric].dropna().to_numpy(dtype=float)
    y_vals = plot_data[y_metric].dropna().to_numpy(dtype=float)
    if x_vals.size and y_vals.size:
        x_min, x_max = x_vals.min(), x_vals.max()
        y_min, y_max = y_vals.min(), y_vals.max()
        x_margin = (x_max - x_min) * 0.08 if x_max > x_min else 0.01
        y_margin = (y_max - y_min) * 0.08 if y_max > y_min else 0.01
        ax2.set_xlim(x_min - x_margin, x_max + x_margin)
        ax2.set_ylim(y_min - y_margin, y_max + y_margin)

    # Calculate correlation
    from scipy.stats import pearsonr

    corr_coef, p_value = pearsonr(plot_data[x_metric], plot_data[y_metric])

    ax2.set_xlabel("Retrieval BEANS Classification R-AUC", fontweight="bold", fontsize=12)
    ax2.set_ylabel("Retrieval BEANS Detection R-AUC", fontweight="bold", fontsize=12)
    ax2.set_title(
        "Bioacoustic generalization separates\nby training paradigm",
        pad=15,
        fontweight="bold",
        fontsize=13,
    )
    ax2.grid(True, alpha=0.25)

    # Create factorial legend: Blue/Orange = SSL/SL, Circle/X = Existing/New
    # Color the legend text to match the marker colors
    legend = ax2.legend(
        loc="upper left",
        fontsize=15,
        title="Training Approach",
        title_fontsize=17,
        frameon=True,
        fancybox=True,
        shadow=True,
    )
    legend.get_title().set_fontweight("bold")

    # Color legend text to match marker colors
    for text in legend.get_texts():
        label_text = text.get_text()
        if label_text in label_to_group:
            group_name = label_to_group[label_text]
            if group_name in ["New SSL Models", "Existing SSL Models"]:
                text.set_color(ssl_color)
            elif group_name in ["New Supervised Models", "Existing Supervised Models"]:
                text.set_color(sl_color)
            else:
                text.set_color("black")  # Post-trained SSL and other groups use black

    # Add panel label
    ax2.text(
        -0.1,
        1.05,
        "(b)",
        transform=ax2.transAxes,
        fontsize=16,
        fontweight="bold",
        va="top",
    )

    # Remove overall figure title - these are separate subfigures
    # fig.suptitle(
    #     "EAT Win-Rate by Benchmark & BEANs Performance Comparison",
    #     fontsize=16, y=0.98
    # )

    plt.tight_layout()

    # Save the plot
    filename = "eat_beans_variant1_by_benchmark.png"
    filepath = output_dir / filename
    plt.savefig(filepath, dpi=300, bbox_inches="tight")
    plt.close()

    print(f"✅ Created variant 1 plot: {filename}")


def create_variant2_plot(df: pd.DataFrame, output_dir: Path) -> None:
    """Create variant 2: Both SSL and Supervised win-rates + BEANs scatter.

    Uses retrieval ROC AUC only.
    """

    # === LEFT SUBPLOT: Win-rates for both SSL and Supervised ===
    wins_dict: Dict[str, Tuple[int, int]] = {}
    avg_dict: Dict[str, float] = {}

    for label, (gen, bio) in BOTH_PAIRS.items():
        imp = improvement_series(df, gen, bio)
        if not imp.empty:
            wins_dict[label] = win_rate(imp)
            avg_dict[label] = imp.mean()

    # === RIGHT SUBPLOT: BEANs Classification vs Detection ===
    benchmark_averages = compute_benchmark_averages(df)

    # Filter to all models that exist in our data
    available_models = [model for model in ALL_MODELS if model in benchmark_averages.index]
    if not available_models:
        print("No models found in data!")
        return

    beans_data = benchmark_averages.loc[available_models]

    # Add model group information for coloring
    model_groups = {}
    for group_name, models in MODEL_GROUPS.items():
        for model in models:
            if model in available_models:
                model_groups[model] = group_name

    beans_data["model_group"] = [model_groups.get(model, "Other") for model in beans_data.index]

    # Metrics to plot - using only retrieval ROC AUC
    x_metric = "BEANS Classification_R-auc"  # Retrieval ROC AUC for BEANS Classification
    y_metric = "BEANS Detection_R-auc"  # Retrieval ROC AUC for BEANS Detection

    # Check if metrics exist
    if x_metric not in beans_data.columns or y_metric not in beans_data.columns:
        print(f"Missing BEANS metrics. Available: {list(beans_data.columns)}")
        return

    # Remove rows with NaN values
    mask = ~(beans_data[x_metric].isna() | beans_data[y_metric].isna())
    plot_data = beans_data[mask]

    if len(plot_data) == 0:
        print("No valid BEANS data for plotting")
        return

    # === CREATE PLOT ===
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 7))

    # LEFT: Win-rates for both model types
    labels = list(wins_dict.keys())
    win_rates = [wins / total * 100 for wins, total in wins_dict.values()]
    counts = [f"{wins}/{total}" for wins, total in wins_dict.values()]
    avg_gains = [avg_dict[label] for label in labels]

    # Shorten labels for better display
    display_labels = []
    for label in labels:
        if "EAT (SSL)" in label:
            display_labels.append("EAT\n(SSL)")
        elif "EffNet (Supervised)" in label:
            display_labels.append("EffNet\n(Supervised)")
        else:
            display_labels.append(label.replace(" ", "\n"))

    # Updated colors for (a) two bars
    # (order matches labels: EAT (SSL), EffNet (Supervised)):
    # SSL -> Blue 4, SL -> Cyan 1
    colors = ["#00738B", "#1ADCCF"]
    bars = ax1.bar(
        display_labels,
        win_rates,
        color=colors,
        alpha=1.0,
        edgecolor="black",
        linewidth=1,
    )

    # Add labels on bars
    for bar, count, avg_gain in zip(bars, counts, avg_gains, strict=False):
        height = bar.get_height()
        ax1.text(
            bar.get_x() + bar.get_width() / 2,
            height + 2,
            f"{height:.1f}%\n{count}\n{avg_gain:+.1f}%",
            ha="center",
            va="bottom",
            fontweight="bold",
            fontsize=15,
        )

    ax1.set_ylabel("Win-Rate (%)", fontweight="bold", fontsize=18)
    ax1.set_title(
        "Benefit of mixing general audio\nin pretraining",
        fontweight="bold",
        pad=15,
        fontsize=20,
    )
    ax1.set_ylim(0, 105)
    ax1.grid(True, alpha=0.3, axis="y")

    # Add panel label
    ax1.text(
        -0.1,
        1.05,
        "(a)",
        transform=ax1.transAxes,
        fontsize=24,
        fontweight="bold",
        va="top",
    )

    # RIGHT: BEANs scatter plot with factorial legend
    # High-contrast palette for (b) to match reference:
    #   SSL -> Blue 4 (#00738B)
    #   SL  -> Cyan 1 (#1ADCCF)
    ssl_color = "#00738B"  # Blue 4 for SSL
    sl_color = "#1ADCCF"  # Cyan 1 for Supervised Learning
    existing_marker = "o"  # Circle
    new_marker = "X"  # Cross
    group_properties = {
        "New Supervised Models": {
            "color": sl_color,
            "marker": new_marker,
            "label": "New SL",
        },
        "Existing Supervised Models": {
            "color": sl_color,
            "marker": existing_marker,
            "label": "Existing SL",
        },
        "New SSL Models": {
            "color": ssl_color,
            "marker": new_marker,
            "label": "New SSL",
        },
        "Existing SSL Models": {
            "color": ssl_color,
            "marker": existing_marker,
            "label": "Existing SSL",
        },
        "Post-trained SSL Models": {
            "color": "black",
            "marker": new_marker,
            "label": "Post-trained SSL",
        },
    }

    # Plot each group with its factorial properties
    plotted_groups = []
    label_to_group = {}  # Map legend labels to group names
    for group_name, props in group_properties.items():
        group_data = plot_data[plot_data["model_group"] == group_name]
        if not group_data.empty:
            ax2.scatter(
                group_data[x_metric],
                group_data[y_metric],
                label=props["label"],
                alpha=1.0,
                s=100,
                edgecolors="k",
                linewidths=0.5,
                color=props["color"],
                marker=props["marker"],
            )
            plotted_groups.append(group_name)
            label_to_group[props["label"]] = group_name

    # Add model labels with better positioning
    # Color labels to match marker colors (blue for SSL, orange for SL, black for others)
    for idx, row in plot_data.iterrows():
        if not pd.isna(row[x_metric]) and not pd.isna(row[y_metric]):
            label = idx
            if "Bird-AVES-biox-base" in label:
                label = "Bird-AVES"

            # Determine label color based on model group
            model_group = row.get("model_group", "Other")
            if model_group in ["New SSL Models", "Existing SSL Models"]:
                label_color = ssl_color
            elif model_group in ["New Supervised Models", "Existing Supervised Models"]:
                label_color = sl_color
            else:
                label_color = "black"  # Post-trained SSL and other groups use black

            ax2.annotate(
                label,
                (row[x_metric], row[y_metric]),
                fontsize=15,
                alpha=0.8,
                ha="center",
                xytext=(2, 2),
                textcoords="offset points",
                color=label_color,
            )

    # Set axis limits with margins
    x_vals = plot_data[x_metric].dropna().to_numpy(dtype=float)
    y_vals = plot_data[y_metric].dropna().to_numpy(dtype=float)
    if x_vals.size and y_vals.size:
        x_min, x_max = x_vals.min(), x_vals.max()
        y_min, y_max = y_vals.min(), y_vals.max()
        x_margin = (x_max - x_min) * 0.08 if x_max > x_min else 0.01
        y_margin = (y_max - y_min) * 0.08 if y_max > y_min else 0.01
        ax2.set_xlim(x_min - x_margin, x_max + x_margin)
        ax2.set_ylim(y_min - y_margin, y_max + y_margin)

    # Calculate correlation
    from scipy.stats import pearsonr

    corr_coef, p_value = pearsonr(plot_data[x_metric], plot_data[y_metric])

    ax2.set_xlabel("Retrieval BEANS Classification R-AUC", fontweight="bold", fontsize=12)
    ax2.set_ylabel("Retrieval BEANS Detection R-AUC", fontweight="bold", fontsize=12)
    ax2.set_title(
        "Bioacoustic generalization separates\nby training paradigm",
        pad=15,
        fontweight="bold",
        fontsize=13,
    )
    ax2.grid(True, alpha=0.25)

    # Create factorial legend: Blue/Orange = SSL/SL, Circle/X = Existing/New
    # Color the legend text to match the marker colors
    legend = ax2.legend(
        loc="upper left",
        fontsize=15,
        title="Training Approach",
        title_fontsize=17,
        frameon=True,
        fancybox=True,
        shadow=True,
    )
    legend.get_title().set_fontweight("bold")

    # Color legend text to match marker colors
    for text in legend.get_texts():
        label_text = text.get_text()
        if label_text in label_to_group:
            group_name = label_to_group[label_text]
            if group_name in ["New SSL Models", "Existing SSL Models"]:
                text.set_color(ssl_color)
            elif group_name in ["New Supervised Models", "Existing Supervised Models"]:
                text.set_color(sl_color)
            else:
                text.set_color("black")  # Post-trained SSL and other groups use black

    # Add panel label
    ax2.text(
        -0.1,
        1.05,
        "(b)",
        transform=ax2.transAxes,
        fontsize=16,
        fontweight="bold",
        va="top",
    )

    # Remove overall figure title - these are separate subfigures
    # fig.suptitle(
    #     "General Audio Training Benefits & BEANs Performance Comparison",
    #     fontsize=16, y=0.98
    # )

    plt.tight_layout()

    # Save the plot
    filename = "eat_beans_variant2_ssl_and_supervised.png"
    filepath = output_dir / filename
    plt.savefig(filepath, dpi=300, bbox_inches="tight")
    plt.close()

    print(f"✅ Created variant 2 plot: {filename}")


def create_variant3_plot(df: pd.DataFrame, output_dir: Path) -> None:
    """Create variant 3: Both SSL and Supervised win-rates.

    Includes CBI vs BEANs Detection scatter.
    """

    # === LEFT SUBPLOT: Win-rates for both SSL and Supervised (same as variant 2) ===
    wins_dict: Dict[str, Tuple[int, int]] = {}
    avg_dict: Dict[str, float] = {}

    for label, (gen, bio) in BOTH_PAIRS.items():
        imp = improvement_series(df, gen, bio)
        if not imp.empty:
            wins_dict[label] = win_rate(imp)
            avg_dict[label] = imp.mean()

    # === RIGHT SUBPLOT: CBI vs BEANS Detection ===
    benchmark_averages = compute_benchmark_averages(df)

    # Filter to all models that exist in our data
    available_models = [model for model in ALL_MODELS if model in benchmark_averages.index]
    if not available_models:
        print("No models found in data!")
        return

    beans_data = benchmark_averages.loc[available_models]

    # Add model group information for coloring
    model_groups = {}
    for group_name, models in MODEL_GROUPS.items():
        for model in models:
            if model in available_models:
                model_groups[model] = group_name

    beans_data["model_group"] = [model_groups.get(model, "Other") for model in beans_data.index]

    # Metrics to plot - CBI dataset specifically vs BEANS Detection
    x_metric = "CBI_R-auc"  # CBI dataset ROC AUC specifically
    y_metric = "BEANS Detection_R-auc"  # Retrieval ROC AUC for BEANS Detection

    # Check if metrics exist - if CBI_R-auc doesn't exist, try to get it from raw data
    if x_metric not in beans_data.columns:
        # Try to get CBI directly from the original dataframe
        if "CBI_R-auc" in df.columns:
            # Add CBI data to beans_data
            for model in available_models:
                if model in df.index:
                    beans_data.loc[model, x_metric] = df.loc[model, "CBI_R-auc"]
        else:
            print(f"CBI dataset not found. Available columns: {list(df.columns)}")
            return

    if y_metric not in beans_data.columns:
        print(f"Missing BEANS Detection metric. Available: {list(beans_data.columns)}")
        return

    # Remove rows with NaN values
    mask = ~(beans_data[x_metric].isna() | beans_data[y_metric].isna())
    plot_data = beans_data[mask]

    if len(plot_data) == 0:
        print("No valid CBI vs BEANS Detection data for plotting")
        return

    # === CREATE PLOT ===
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 7))

    # LEFT: Win-rates for both model types (same as variant 2)
    labels = list(wins_dict.keys())
    win_rates = [wins / total * 100 for wins, total in wins_dict.values()]
    counts = [f"{wins}/{total}" for wins, total in wins_dict.values()]
    avg_gains = [avg_dict[label] for label in labels]

    # Shorten labels for better display
    display_labels = []
    for label in labels:
        if "EAT (SSL)" in label:
            display_labels.append("EAT\n(SSL)")
        elif "EffNet (Supervised)" in label:
            display_labels.append("EffNet\n(Supervised)")
        else:
            display_labels.append(label.replace(" ", "\n"))

    colors = ["#4ECDC4", "#FF6B6B"]  # Teal for SSL, Red for Supervised
    bars = ax1.bar(
        display_labels,
        win_rates,
        color=colors,
        alpha=0.8,
        edgecolor="black",
        linewidth=1,
    )

    # Add labels on bars
    for bar, count, avg_gain in zip(bars, counts, avg_gains, strict=False):
        height = bar.get_height()
        ax1.text(
            bar.get_x() + bar.get_width() / 2,
            height + 2,
            f"{height:.1f}%\n{count}\n{avg_gain:+.1f}%",
            ha="center",
            va="bottom",
            fontweight="bold",
            fontsize=15,
        )

    ax1.set_ylabel("Win-Rate (%)", fontweight="bold", fontsize=18)
    ax1.set_title(
        "Benefit of mixing general audio\nin pretraining",
        fontweight="bold",
        pad=15,
        fontsize=20,
    )
    ax1.set_ylim(0, 105)
    ax1.grid(True, alpha=0.3, axis="y")

    # Add panel label
    ax1.text(
        -0.1,
        1.05,
        "(a)",
        transform=ax1.transAxes,
        fontsize=24,
        fontweight="bold",
        va="top",
    )

    # RIGHT: CBI vs BEANS Detection scatter plot with factorial legend
    # Factorial design: Color = SSL/SL, Shape = New/Existing

    # Colorblind-safe colors: Blue for SSL, Cyan for SL
    ssl_color = "#00738B"  # Blue 4 for SSL
    sl_color = "#1ADCCF"  # Cyan 1 for Supervised Learning

    # Markers: circle for existing, X for new
    existing_marker = "o"  # Circle
    new_marker = "X"  # Cross

    # Define groups and their properties
    group_properties = {
        "New Supervised Models": {
            "color": sl_color,
            "marker": new_marker,
            "label": "New SL",
        },
        "Existing Supervised Models": {
            "color": sl_color,
            "marker": existing_marker,
            "label": "Existing SL",
        },
        "New SSL Models": {
            "color": ssl_color,
            "marker": new_marker,
            "label": "New SSL",
        },
        "Existing SSL Models": {
            "color": ssl_color,
            "marker": existing_marker,
            "label": "Existing SSL",
        },
        "Post-trained SSL Models": {
            "color": "black",
            "marker": new_marker,
            "label": "Post-trained SSL",
        },
    }

    # Plot each group with its factorial properties
    plotted_groups = []
    label_to_group = {}  # Map legend labels to group names
    for group_name, props in group_properties.items():
        group_data = plot_data[plot_data["model_group"] == group_name]
        if not group_data.empty:
            ax2.scatter(
                group_data[x_metric],
                group_data[y_metric],
                label=props["label"],
                alpha=0.8,
                s=100,
                edgecolors="k",
                linewidths=0.5,
                color=props["color"],
                marker=props["marker"],
            )
            plotted_groups.append(group_name)
            label_to_group[props["label"]] = group_name

    # Add model labels with better positioning
    # Color labels to match marker colors (blue for SSL, orange for SL, black for others)
    for idx, row in plot_data.iterrows():
        if not pd.isna(row[x_metric]) and not pd.isna(row[y_metric]):
            label = idx
            if "Bird-AVES-biox-base" in label:
                label = "Bird-AVES"

            # Determine label color based on model group
            model_group = row.get("model_group", "Other")
            if model_group in ["New SSL Models", "Existing SSL Models"]:
                label_color = ssl_color
            elif model_group in ["New Supervised Models", "Existing Supervised Models"]:
                label_color = sl_color
            else:
                label_color = "black"  # Post-trained SSL and other groups use black

            ax2.annotate(
                label,
                (row[x_metric], row[y_metric]),
                fontsize=15,
                alpha=0.8,
                ha="center",
                xytext=(2, 2),
                textcoords="offset points",
                color=label_color,
            )

    # Set axis limits with margins
    x_vals = plot_data[x_metric].dropna().to_numpy(dtype=float)
    y_vals = plot_data[y_metric].dropna().to_numpy(dtype=float)
    if x_vals.size and y_vals.size:
        x_min, x_max = x_vals.min(), x_vals.max()
        y_min, y_max = y_vals.min(), y_vals.max()
        x_margin = (x_max - x_min) * 0.08 if x_max > x_min else 0.01
        y_margin = (y_max - y_min) * 0.08 if y_max > y_min else 0.01
        ax2.set_xlim(x_min - x_margin, x_max + x_margin)
        ax2.set_ylim(y_min - y_margin, y_max + y_margin)

    # Calculate correlation
    from scipy.stats import pearsonr

    corr_coef, p_value = pearsonr(plot_data[x_metric], plot_data[y_metric])

    ax2.set_xlabel("CBI Dataset ROC-AUC", fontweight="bold", fontsize=12)
    ax2.set_ylabel("Retrieval BEANS Detection R-AUC", fontweight="bold", fontsize=12)
    ax2.set_title(
        "Performance on CBI vs BEANS Detection\nseparates by training paradigm",
        pad=15,
        fontweight="bold",
        fontsize=13,
    )
    ax2.grid(True, alpha=0.25)

    # Create factorial legend: Blue/Orange = SSL/SL, Circle/X = Existing/New
    # Color the legend text to match the marker colors
    legend = ax2.legend(
        loc="upper left",
        fontsize=15,
        title="Training Approach",
        title_fontsize=17,
        frameon=True,
        fancybox=True,
        shadow=True,
    )
    legend.get_title().set_fontweight("bold")

    # Color legend text to match marker colors
    for text in legend.get_texts():
        label_text = text.get_text()
        if label_text in label_to_group:
            group_name = label_to_group[label_text]
            if group_name in ["New SSL Models", "Existing SSL Models"]:
                text.set_color(ssl_color)
            elif group_name in ["New Supervised Models", "Existing Supervised Models"]:
                text.set_color(sl_color)
            else:
                text.set_color("black")  # Post-trained SSL and other groups use black

    # Add panel label
    ax2.text(
        -0.1,
        1.05,
        "(b)",
        transform=ax2.transAxes,
        fontsize=16,
        fontweight="bold",
        va="top",
    )

    plt.tight_layout()

    # Save the plot
    filename = "eat_beans_variant3_cbi_specific.png"
    filepath = output_dir / filename
    plt.savefig(filepath, dpi=300, bbox_inches="tight")
    plt.close()

    print(f"✅ Created variant 3 plot: {filename}")


def main() -> None:
    """Main execution function."""
    print(
        "🔬 Creating EAT + BEANS Combined Analysis Variants "
        "(three variants: retrieval aggregates, retrieval aggregates, "
        "and CBI-specific)..."
    )

    # Load data
    df = load_df(EXCEL)

    # Create output directory
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = OUT_BASE / f"{ts}_eat_beans_variants"
    out_dir.mkdir(parents=True, exist_ok=True)

    # Create all three variants
    create_variant1_plot(df, out_dir)
    create_variant2_plot(df, out_dir)
    create_variant3_plot(df, out_dir)

    print(f"✅ Analysis complete! Results saved to: {out_dir}")


if __name__ == "__main__":
    main()
